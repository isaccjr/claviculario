
from datetime import timedelta
from datetime import datetime
import traceback

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.utils import timezone
from django.contrib.auth.models import User # Importa o modelo User
from django.views.generic import ListView, CreateView, UpdateView # Importa as CBVs do Django

# Imports para Geração de Relatórios
import csv
from openpyxl import Workbook
import pandas as pd
from django.db.models import Count, F

from django.contrib.auth.mixins import AccessMixin
# IMPORTAÇÃO DOS NOSSOS MIXINS CUSTOMIZADOS
from .mixins import (
    LoginRequiredMixin,PaginaAtivaMixin, BaseListView, BaseCreateView, BaseUpdateView, BaseDesativarView,
    BasePessoaView, BaseChaveView, BaseLocalView
)

# Imports dos Modelos e Formulários
from .models import Emprestimo, Chave, Pessoa, Local
from .forms import (
    EmprestimoForm, RelatorioForm, PessoaForm, ChaveForm, LocalForm,
    CustomUserCreationForm, CustomUserChangeForm # Importa os novos formulários de usuário
)

@login_required
def view_retirada(request):
    if request.method == 'POST':
        form = EmprestimoForm(request.POST)
        if form.is_valid():
            chave_selecionada = form.cleaned_data['chave']
            if not chave_selecionada.disponivel:
                messages.error(request, 'Esta chave já foi emprestada. Por favor, selecione outra.')
                return redirect('view_retirada')
            form.save()
            chave_selecionada.disponivel = False
            chave_selecionada.save()
            messages.success(request, f'Chave "{chave_selecionada.descricao}" emprestada com sucesso!')
            return redirect('view_retirada')
    else:
        form = EmprestimoForm()

    emprestimos_ativos = Emprestimo.objects.filter(data_devolucao__isnull=True).order_by('-data_retirada')
    form_pessoa = PessoaForm()
    locais = Local.objects.all().order_by('nome')
    contexto = {
        'form': form,
        'form_pessoa' : form_pessoa,
        'emprestimos_ativos': emprestimos_ativos,
        'locais' : locais,
        'pagina_ativa': 'retirada' # Para o menu lateral
    }
    return render(request, 'claviculario_app/retirada.html', contexto)

# NOVA VIEW PARA A PÁGINA DE DEVOLUÇÃO
@login_required
def view_devolucao(request):
    emprestimos_ativos = Emprestimo.objects.filter(data_devolucao__isnull=True).order_by('chave__descricao')
    
    # Filtra as chaves e pessoas que têm empréstimos ativos
    chaves_emprestadas = Chave.objects.filter(id__in=emprestimos_ativos.values_list('chave_id', flat=True)).distinct()
    pessoas_com_chave = Pessoa.objects.filter(id__in=emprestimos_ativos.values_list('pessoa_id', flat=True)).distinct()

    # Lógica para filtros GET
    chave_id_filtrada = request.GET.get('chave')
    pessoa_id_filtrada = request.GET.get('pessoa')

    if chave_id_filtrada:
        emprestimos_ativos = emprestimos_ativos.filter(chave_id=chave_id_filtrada)
    
    if pessoa_id_filtrada:
        emprestimos_ativos = emprestimos_ativos.filter(pessoa_id=pessoa_id_filtrada)

    contexto = {
        'emprestimos': emprestimos_ativos,
        'chaves_emprestadas': chaves_emprestadas,
        'pessoas_com_chave': pessoas_com_chave,
        'pagina_ativa': 'devolucao' # Para o menu lateral
    }
    return render(request, 'claviculario_app/devolucao.html', contexto)

def paginador(request,lista:list,quant_por_pag:int = 15) :
    # Cria um Paginator com a nossa lista , mostrando n itens por página
    paginator = Paginator(lista, quant_por_pag)
    
    # Pega o número da página da URL (ex: ?page=2)
    page_number = request.GET.get('page')
    
    # Pega o objeto da página correta
    pagina = paginator.get_page(page_number)

    return pagina

# NOVA VIEW PARA A PÁGINA DE RELATÓRIO
@login_required
def view_relatorio(request):
    form = RelatorioForm(request.GET)
    emprestimos_list = _get_emprestimos_filtrados(request)
    contexto = {
        'form': form,
        'emprestimos_page': paginador(request,emprestimos_list),
        'pagina_ativa': 'relatorio' # Para o menu lateral
    }
    return render(request, 'claviculario_app/relatorio.html', contexto)

# Esta função permanece quase a mesma, apenas o redirect muda
@login_required
def registrar_devolucao(request, emprestimo_id):
    if request.method == 'POST':
        emprestimo = get_object_or_404(Emprestimo, id=emprestimo_id)
        if emprestimo.data_devolucao is not None:
            messages.warning(request, 'Esta chave já foi devolvida anteriormente.')
        else:
            emprestimo.data_devolucao = timezone.now()
            emprestimo.save()
            chave = emprestimo.chave
            chave.disponivel = True
            chave.save()
            messages.info(request, f'Devolução da chave "{chave.descricao}" registrada com sucesso.')
    
    # Redireciona de volta para a página de devolução
    return redirect('view_devolucao')

# NOVA VIEW PARA CADASTRAR PESSOA
@login_required
def cadastrar_pessoa(request):
    # Dicionário para a resposta JSON
    data = {'success': False}
    if request.method == 'POST':
        form = PessoaForm(request.POST)
        if form.is_valid():
            pessoa = form.save()
            data['success'] = True
            # Retorna os dados da nova pessoa para adicionar dinamicamente ao select
            data['pessoa'] = {'id': pessoa.id, 'nome': str(pessoa)}
        else:
            # Coleta os erros do formulário para exibir no modal
            data['errors'] = form.errors.as_json()
    
    # Retorna uma resposta JSON em vez de renderizar um template
    return JsonResponse(data)

#  VIEW PARA SERVIR COMO API DE FILTRO
@login_required
def filtrar_pessoas(request):
    # Pega os parâmetros da URL (ex: /api/pessoas/?nome=joao&empresa=fab)
    nome_query = request.GET.get('nome', '')
    empresa_query = request.GET.get('empresa', '')

    pessoas = Pessoa.objects.all()

    if nome_query:
        # Filtra por nome (case-insensitive)
        pessoas = pessoas.filter(nome__icontains=nome_query)
    
    if empresa_query:
        # Filtra por empresa (case-insensitive)
        pessoas = pessoas.filter(empresa__icontains=empresa_query)

    # Ordena por nome
    pessoas = pessoas.order_by('nome')

    # Transforma a lista de objetos Pessoa em um formato simples (JSON)
    data = [{'id': p.id, 'text': str(p)} for p in pessoas]

    return JsonResponse({'results': data})


#VIEW PARA VALIDAR O PIN E CONCLUIR A RETIRADA
@login_required
def verificar_pin_e_registrar(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método inválido.'})

    try:
        chave_id = request.POST.get('chave_id')
        pessoa_id = request.POST.get('pessoa_id')
        pin_digitado = request.POST.get('pin')
        observacao = request.POST.get('observacao')
        data_retirada = request.POST.get('data_retirada')
        
        # 1. ADICIONADO: Capturar a data de previsão do formulário
        # O 'or None' garante que se o campo vier vazio, salvamos como nulo no banco
        previsao_devolucao = request.POST.get('previsao_devolucao') or None

        pessoa = Pessoa.objects.get(pk=pessoa_id)
        chave = Chave.objects.get(pk=chave_id)

        if not pessoa.check_pin(pin_digitado):
            return JsonResponse({'success': False, 'message': 'PIN incorreto!'})

        if not chave.disponivel:
            return JsonResponse({'success': False, 'message': 'Esta chave foi retirada por outra pessoa. Atualize a página.'})
            
        Emprestimo.objects.create(
            chave=chave,
            pessoa=pessoa,
            data_retirada=data_retirada,
            observacao=observacao,
            # 2. ADICIONADO: Passar o valor para ser salvo no banco de dados
            previsao_devolucao=previsao_devolucao
        )

        chave.disponivel = False
        chave.save()

        return JsonResponse({'success': True, 'message': f'Chave "{chave.descricao}" emprestada com sucesso!'})

    except Pessoa.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Pessoa não encontrada.'})
    except Chave.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Chave não encontrada.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Ocorreu um erro: {e}'})


#VIEW DO DASHBOARD (FAZ CALCULOS DE CHAVES)
@login_required
def dashboard(request):
    # Cálculos dos cards
    total_chaves = Chave.objects.count()
    chaves_emprestadas = Emprestimo.objects.filter(data_devolucao__isnull=True)
    chaves_emprestadas_count = chaves_emprestadas.count()
    chaves_disponiveis = total_chaves - chaves_emprestadas_count
    
    # O cálculo de chaves atrasadas agora é possível
    chaves_atrasadas_count = chaves_emprestadas.filter(
        previsao_devolucao__isnull=False, 
        previsao_devolucao__lt=timezone.now()
    ).count()

    # Listas para a Dashboard
    ultimas_atividades = Emprestimo.objects.order_by('-data_retirada')[:5] # Pega as 5 últimas retiradas
    emprestimos_atrasados = chaves_emprestadas.filter(
        previsao_devolucao__isnull=False,
        previsao_devolucao__lt=timezone.now()
    ).order_by('previsao_devolucao') # Pega todas as chaves atrasadas, ordenadas pela mais antiga

    contexto = {
        'total_chaves': total_chaves,
        'chaves_emprestadas_count': chaves_emprestadas_count,
        'chaves_disponiveis': chaves_disponiveis,
        'chaves_atrasadas_count': chaves_atrasadas_count,
        'ultimas_atividades': ultimas_atividades,
        'emprestimos_atrasados': emprestimos_atrasados,
        'pagina_ativa': 'dashboard' # Para o menu lateral
    }
    return render(request, 'claviculario_app/dashboard.html', contexto)


#------------------------------------------------------------------
# VIEWS DE PESSOAS
#------------------------------------------------------------------
#VIEW PARA SERVIR COMO API DE FILTRO DE CHAVES


# --- CLASS-BASED VIEWS PARA O CRUD DE PESSOAS ---

class PessoaListView(BasePessoaView, BaseListView):
    queryset = BasePessoaView.model.objects.filter(ativa=True).order_by('nome')

class PessoaCreateView(BasePessoaView, BaseCreateView):
    title = 'Adicionar Nova Pessoa'

class PessoaUpdateView(BasePessoaView, BaseUpdateView):
    title = 'Editar Pessoa: {objeto.nome}'

class PessoaDesativarView(BaseDesativarView):
    model = Pessoa
    pagina_ativa = 'pessoas'
    success_message = "A Pessoa '{objeto.nome}' foi desativada com sucesso."
    def form_valid(self, form):
        pessoa = self.get_object()
        if pessoa.emprestimos.filter(data_devolucao__isnull=True).exists():
            messages.error(self.request, f"A pessoa '{pessoa.nome}' não pode ser desativada pois possui chaves pendentes.")
            return redirect('pessoa_list')
        return super().form_valid(form)

@login_required
def filtrar_chaves_por_local(request):
    local_id = request.GET.get('local_id')
    chaves = Chave.objects.filter(disponivel=True, ativa=True) # Começa com todas as chaves disponíveis

    if local_id:
        chaves = chaves.filter(local_id=local_id)
    else:
        # Se nenhum local for selecionado, retorna uma lista vazia
        chaves = Chave.objects.none()

    chaves = chaves.order_by('descricao')
    
    # Transforma a lista de objetos Chave em um formato simples (JSON)
    data = [{'id': c.id, 'text': str(c)} for c in chaves]

    return JsonResponse({'results': data})

@permission_required('claviculario_app.view_pessoa', raise_exception=True)
def pessoa_historico(request, pk):
    pessoa = get_object_or_404(Pessoa, pk=pk)
    # Busca todos os empréstimos dessa pessoa, ordenando pelos mais recentes
    emprestimos_list = pessoa.emprestimos.select_related('chave', 'chave__local').order_by('-data_retirada')    
    contexto = {
        'pessoa': pessoa,
        'emprestimos_page': paginador(request,emprestimos_list),
        'pagina_ativa': 'pessoas'
    }
    return render(request, 'claviculario_app/pessoa_historico.html', contexto)


#------------------------------------------------------------------
# VIEWS DE CHAVES
#------------------------------------------------------------------
class ChaveListView(BaseChaveView, BaseListView):
    queryset = BaseChaveView.model.objects.filter(ativa=True).select_related('local').order_by('local__nome', 'descricao')

class ChaveCreateView(BaseChaveView, BaseCreateView):
    title = 'Adicionar Nova Chave'

class ChaveUpdateView(BaseChaveView, BaseUpdateView):
    title = 'Editar Chave: {objeto.descricao}'

class ChaveDesativarView(BaseDesativarView):
    model = Chave # <-- Definimos o model explicitamente
    pagina_ativa = 'chaves'
    success_message = "A Chave '{objeto.descricao}' foi desativada com sucesso."
    def form_valid(self, form):
        chave = self.get_object()
        if not chave.disponivel:
            messages.error(self.request, f"A chave '{chave.descricao}' não pode ser desativada pois está emprestada.")
            return redirect('chave_list')
        return super().form_valid(form)

@permission_required('claviculario_app.view_chave', raise_exception=True)
def chave_historico(request, pk):
    chave = get_object_or_404(Chave, pk=pk)
    emprestimos_list = chave.emprestimos.select_related('pessoa').order_by('-data_retirada')
    # Aplicamos a paginação
    paginator = Paginator(emprestimos_list, 15)
    page_number = request.GET.get('page')
    emprestimos_page = paginator.get_page(page_number)
    
    contexto = {
        'chave': chave,
        'emprestimos_page': paginador(request,emprestimos_list),
        'pagina_ativa': 'chaves'
    }
    return render(request, 'claviculario_app/chave_historico.html', contexto)

#------------------------------------------------------------------
# VIEWS DE LOCAIS
#------------------------------------------------------------------

class LocalListView(BaseLocalView, BaseListView):
    queryset = BaseLocalView.model.objects.filter(ativa=True).order_by('nome')

class LocalCreateView(BaseLocalView, BaseCreateView):
    title = 'Adicionar Novo Local'

class LocalUpdateView(BaseLocalView, BaseUpdateView):
    title = 'Editar Local: {objeto.nome}'

class LocalDesativarView(BaseDesativarView):
    model = Local # <-- Definimos o model explicitamente
    pagina_ativa = 'locais'
    success_message = "O Local '{objeto.nome}' foi desativado com sucesso."
    def form_valid(self, form):
        local = self.get_object()
        if local.chaves.filter(ativa=True).exists():
            messages.error(self.request, f"O local '{local.nome}' não pode ser desativado pois ainda existem chaves ativas associadas a ele.")
            return redirect('local_list')
        return super().form_valid(form)
# LÓGICA DE FILTRO REUTILIZÁVEL (FUNÇÃO AUXILIAR)
# claviculario_app/views.py

def _get_emprestimos_filtrados(request):
    print("\n--- INICIANDO DEBUG DO FILTRO DE RELATÓRIO ---")
    print(f"[PASSO 1] Dados recebidos na URL (request.GET): {request.GET}")

    form = RelatorioForm(request.GET)
    emprestimos_list = Emprestimo.objects.select_related('chave', 'pessoa').order_by('-data_retirada')

    print(f"[PASSO 2] O formulário é válido? {form.is_valid()}")

    if form.is_valid():
        print(f"[PASSO 3] Dados 'limpos' do formulário (cleaned_data): {form.cleaned_data}")
        
        data_inicio = form.cleaned_data.get('data_inicio')
        data_fim = form.cleaned_data.get('data_fim')
        status = form.cleaned_data.get('status')
        pessoa = form.cleaned_data.get('pessoa')
        chave = form.cleaned_data.get('chave')

        if data_inicio:
            emprestimos_list = emprestimos_list.filter(data_retirada__date__gte=data_inicio)
        if data_fim:
            emprestimos_list = emprestimos_list.filter(data_retirada__date__lte=data_fim)
        if status == 'pendentes':
            emprestimos_list = emprestimos_list.filter(data_devolucao__isnull=True)
            
        if pessoa:
            print(f"-> Aplicando filtro para a pessoa: {pessoa}")
            emprestimos_list = emprestimos_list.filter(pessoa=pessoa)
        if chave:
            print(f"-> Aplicando filtro para a chave: {chave}")
            emprestimos_list = emprestimos_list.filter(chave=chave)
    else:
        # Se o formulário não for válido, vamos ver os erros
        print(f"[ERRO] Erros de validação do formulário: {form.errors}")
    
    print("--- FIM DO DEBUG ---\n")
    return emprestimos_list

# NOVA VIEW PARA EXPORTAR CSV
@login_required
def exportar_relatorio_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="relatorio_claviculario.csv"'
    
    emprestimos = _get_emprestimos_filtrados(request)
    
    writer = csv.writer(response)
    # Cabeçalho do CSV
    writer.writerow(['Chave', 'Responsável', 'CPF/SARAN', 'Data Retirada', 'Data Devolução', 'Observação'])
    
    # Linhas
    for emprestimo in emprestimos:
        writer.writerow([
            emprestimo.chave.descricao,
            emprestimo.pessoa.nome,
            emprestimo.pessoa.cpf_saran,
            emprestimo.data_retirada.strftime('%d/%m/%Y %H:%M'),
            emprestimo.data_devolucao.strftime('%d/%m/%Y %H:%M') if emprestimo.data_devolucao else 'Pendente',
            emprestimo.observacao
        ])
        
    return response

# VIEW PARA EXPORTAR EXCEL
@login_required
def exportar_relatorio_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_claviculario.xlsx"'
    
    emprestimos = _get_emprestimos_filtrados(request)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    
    headers = ['Chave', 'Responsável', 'CPF/SARAN', 'Data Retirada', 'Data Devolução', 'Observação']
    ws.append(headers)
    

    for emprestimo in emprestimos:
        # Convert aware datetimes to local time
        data_retirada_local = timezone.localtime(emprestimo.data_retirada)
        
        # Make the datetime "naive" (remove timezone info) for Excel
        data_retirada_naive = data_retirada_local.replace(tzinfo=None)
        
        data_devolucao_final = 'Pendente'
        if emprestimo.data_devolucao:
            data_devolucao_local = timezone.localtime(emprestimo.data_devolucao)
            data_devolucao_final = data_devolucao_local.replace(tzinfo=None)
            
        ws.append([
            emprestimo.chave.descricao,
            emprestimo.pessoa.nome,
            emprestimo.pessoa.cpf_saran,
            data_retirada_naive, # Use the naive datetime
            data_devolucao_final, # Use the naive datetime or 'Pendente'
            emprestimo.observacao
        ])
        
    wb.save(response)
    return response


# VIEW PARA A PÁGINA DE IMPORTAÇÃO
@permission_required('claviculario_app.add_pessoa', raise_exception=True) # Só gerentes podem importar
def importar_dados_page(request):
    contexto = {
        'pagina_ativa': 'importar'
    }
    return render(request, 'claviculario_app/importar_dados_page.html', contexto)


# VIEW PARA GERAR E FAZER O DOWNLOAD DO TEMPLATE DE PESSOAS
@permission_required('claviculario_app.add_pessoa', raise_exception=True)
def download_template_pessoas(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="template_pessoas.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pessoas"
    
    # Define os cabeçalhos (nomes de coluna sem espaços ou acentos para facilitar)
    headers = ['nome_completo', 'empresa', 'cpf_saran', 'pin']
    ws.append(headers)
    
    wb.save(response)
    return response

# VIEW PARA GERAR E FAZER O DOWNLOAD DO TEMPLATE DE CHAVES
@permission_required('claviculario_app.add_chave', raise_exception=True)
def download_template_chaves(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="template_chaves.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Chaves"
    
    headers = ['descricao_chave', 'nome_local']
    ws.append(headers)
    
    wb.save(response)
    return response


@permission_required('claviculario_app.add_pessoa', raise_exception=True)
def importar_pessoas(request):
    if request.method == 'POST':
        arquivo = request.FILES.get('arquivo_excel')
        if not arquivo or not arquivo.name.endswith('.xlsx'):
            messages.error(request, "Por favor, envie um arquivo Excel (.xlsx) válido.")
            return redirect('importar_dados_page')

        try:
            df = pd.read_excel(arquivo)
            # Verifica se as colunas necessárias existem
            colunas_necessarias = ['nome_completo', 'cpf_saran', 'pin']
            if not all(coluna in df.columns for coluna in colunas_necessarias):
                messages.error(request, f"O arquivo enviado não contém as colunas necessárias: {colunas_necessarias}")
                return redirect('importar_dados_page')

            criados, ignorados, conflitos = 0, 0, []

            for index, row in df.iterrows():
                cpf_saran = str(row['cpf_saran']).strip()
                nome = str(row['nome_completo']).strip()
                pin = str(row['pin']).strip()

                if not cpf_saran or not nome or not pin:
                    continue # Ignora linhas com dados essenciais faltando

                pessoa_existente = Pessoa.objects.filter(cpf_saran=cpf_saran).first()

                if pessoa_existente:
                    if pessoa_existente.nome != nome:
                        conflitos.append(f"CPF/SARAN {cpf_saran}: Nome no sistema '{pessoa_existente.nome}', nome na planilha '{nome}'. Nenhum dado foi alterado.")
                    ignorados += 1
                else:
                    nova_pessoa = Pessoa(
                        nome=nome,
                        cpf_saran=cpf_saran,
                        empresa=row.get('empresa'),
                    )
                    nova_pessoa.set_pin(pin)
                    nova_pessoa.save()
                    criados += 1
            
            messages.success(request, f"Importação de pessoas concluída! {criados} pessoas criadas, {ignorados} já existentes ignoradas.")
            if conflitos:
                messages.warning(request, f"Atenção: {len(conflitos)} registros apresentaram conflito de nome e não foram atualizados. Verifique os dados.")
                # Opcional: mostrar os detalhes dos conflitos
                # for conflito in conflitos:
                #     messages.info(request, conflito)

        except Exception as e:
            messages.error(request, f"Ocorreu um erro ao processar o arquivo: {e}")

    return redirect('importar_dados_page')


# NOVA VIEW PARA PROCESSAR A IMPORTAÇÃO DE CHAVES
@permission_required('claviculario_app.add_chave', raise_exception=True)
def importar_chaves(request):
    if request.method == 'POST':
        arquivo = request.FILES.get('arquivo_excel')
        if not arquivo or not arquivo.name.endswith('.xlsx'):
            messages.error(request, "Por favor, envie um arquivo Excel (.xlsx) válido.")
            return redirect('importar_dados_page')

        try:
            df = pd.read_excel(arquivo)
            colunas_necessarias = ['descricao_chave', 'nome_local']
            if not all(coluna in df.columns for coluna in colunas_necessarias):
                messages.error(request, f"O arquivo enviado não contém as colunas necessárias: {colunas_necessarias}")
                return redirect('importar_dados_page')

            criadas, ignoradas = 0, 0
            
            for index, row in df.iterrows():
                descricao = str(row['descricao_chave']).strip()
                nome_local = str(row['nome_local']).strip()

                if not descricao or not nome_local:
                    continue

                if Chave.objects.filter(descricao=descricao).exists():
                    ignoradas += 1
                    continue
                
                # Procura pelo local, se não existir, cria um novo (get_or_create)
                local, local_criado = Local.objects.get_or_create(
                    nome=nome_local,
                    defaults={'ativa': True} # Define 'ativa' como True se criar um novo
                )
                
                Chave.objects.create(descricao=descricao, local=local)
                criadas += 1
            
            messages.success(request, f"Importação de chaves concluída! {criadas} chaves criadas, {ignoradas} já existentes ignoradas.")

        except Exception as e:
            messages.error(request, f"Ocorreu um erro ao processar o arquivo: {e}")

    return redirect('importar_dados_page')


@login_required
def analytics_data(request):
    try:
        # --- 1. CAPTURAR E VALIDAR FILTROS ---
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        group_by = request.GET.get('group_by', 'day')
        local_id = request.GET.get('local_id')
        chave_id = request.GET.get('chave_id')

        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else timezone.now().date()
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else end_date - timedelta(days=29)

        # --- 2. FILTRAR O CONJUNTO DE DADOS PRINCIPAL ---
        queryset = Emprestimo.objects.filter(data_retirada__date__gte=start_date, data_retirada__date__lte=end_date)
        if local_id: queryset = queryset.filter(chave__local_id=local_id)
        if chave_id: queryset = queryset.filter(chave_id=chave_id)

        # --- 3. GRÁFICO 1: ATRASOS ---
        devolvidos_qs = queryset.filter(data_devolucao__isnull=False, previsao_devolucao__isnull=False)
        atrasados = devolvidos_qs.filter(data_devolucao__gt=F('previsao_devolucao')).count()
        em_dia = devolvidos_qs.count() - atrasados
        dados_atrasos = {'labels': ['Em Dia', 'Com Atraso'], 'data': [em_dia, atrasados]}

        # --- 4. FUNÇÃO AUXILIAR PARA PROCESSAR OS GRÁFICOS DE SÉRIE ---
        def processar_agrupamento(qs, date_field):
            if not qs.exists(): return [], []
            datas = [timezone.localtime(dt) for dt in qs.values_list(date_field, flat=True) if dt]
            if not datas: return [], []
            
            df = pd.DataFrame(datas, columns=['datetime'])
            num_dias_no_periodo = (end_date - start_date).days + 1
            
            if group_by in ['time_of_day', 'time_of_day_avg']:
                counts = df['datetime'].dt.hour.value_counts()
                labels = [f"{h:02d}:00" for h in range(24)]
                data_total = [counts.get(h, 0) for h in range(24)]
                data = [total / num_dias_no_periodo for total in data_total] if group_by == 'time_of_day_avg' else data_total
                return labels, data
            elif group_by == 'day':
                counts = df.set_index('datetime').resample('D').size()
                date_range = pd.date_range(start=start_date, end=end_date, freq='D', tz='America/Sao_Paulo')
                counts = counts.reindex(date_range, fill_value=0)
                labels = counts.index.strftime('%d/%m/%Y').tolist()
                data = counts.values.tolist()
                return labels, data
            elif group_by in ['weekday', 'weekday_avg']:
                counts = df['datetime'].dt.weekday.value_counts()
                weekday_map = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']
                data_total = [counts.get(i, 0) for i in range(7)]
                if group_by == 'weekday_avg':
                    num_semanas = max(1, num_dias_no_periodo / 7.0)
                    data = [total / num_semanas for total in data_total]
                else:
                    data = data_total
                return weekday_map, data
            elif group_by in ['monthday', 'monthday_avg']:
                counts = df['datetime'].dt.day.value_counts()
                labels = [str(i) for i in range(1, 32)]
                data_total = [counts.get(i, 0) for i in range(1, 32)]
                if group_by == 'monthday_avg':
                    num_meses = max(1, (end_date.year - start_date.year) * 12 + end_date.month - start_date.month + 1)
                    data = [total / num_meses for total in data_total]
                else:
                    data = data_total
                return labels, data
            return [], []

        # --- 5. PROCESSAR E RETORNAR OS DADOS FINAIS ---
        retiradas_labels, retiradas_data = processar_agrupamento(queryset, 'data_retirada')
        devolucoes_labels, devolucoes_data = processar_agrupamento(queryset.filter(data_devolucao__isnull=False), 'data_devolucao')

        # --- A CORREÇÃO ESTÁ AQUI ---
        # Converte os números do tipo do pandas (int64) para float padrão do Python
        retiradas_data = [float(x) for x in retiradas_data]
        devolucoes_data = [float(x) for x in devolucoes_data]
        
        labels_finais = retiradas_labels if len(retiradas_labels) >= len(devolucoes_labels) else devolucoes_labels
        if len(devolucoes_data) < len(labels_finais):
            devolucoes_data.extend([0] * (len(labels_finais) - len(devolucoes_data)))

        data = {
            'atrasos': dados_atrasos,
            'retiradas': {'labels': labels_finais, 'data': retiradas_data},
            'devolucoes': {'labels': labels_finais, 'data': devolucoes_data},
        }
        return JsonResponse(data)

    except Exception as e:
        error_trace = traceback.format_exc()
        print("="*50, "\nERRO GRAVE NA VIEW analytics_data:\n", error_trace, "\n"+"="*50)
        return JsonResponse({'error': f"Erro no servidor: {e}", 'traceback': error_trace}, status=500)

@login_required
# Só gerentes podem ver a página de análise
@permission_required('claviculario_app.view_emprestimo', raise_exception=True)
def analytics_page(request):
    # Precisamos enviar a lista de chaves e locais para os filtros
    chaves = Chave.objects.filter(ativa=True).order_by('descricao')
    locais = Local.objects.filter(ativa=True).order_by('nome')
    
    contexto = {
        'chaves': chaves,
        'locais': locais,
        'pagina_ativa': 'analise'
    }
    return render(request, 'claviculario_app/analytics_page.html', contexto)



# --- CRUD de Contas de Usuário ---
# claviculario_app/views.py

class UserListView(LoginRequiredMixin, PaginaAtivaMixin, ListView):
    model = User
    template_name = 'claviculario_app/user_list.html'
    context_object_name = 'usuarios_page'
    paginate_by = 15
    pagina_ativa = 'contas'
    
    def get_queryset(self):
        if self.request.user.is_superuser:
            # CORREÇÃO AQUI: Adiciona o filtro para mostrar apenas usuários ativos.
            return User.objects.filter(is_active=True).order_by('username')
        return User.objects.none()

class UserCreateView(PaginaAtivaMixin, LoginRequiredMixin, CreateView):
    model = User
    form_class = CustomUserCreationForm
    template_name = 'claviculario_app/user_form.html'
    success_url = reverse_lazy('user_list')
    pagina_ativa = 'contas'
    title = 'Criar Novo Usuário'
    def form_valid(self, form):
        messages.success(self.request, 'Usuário criado com sucesso!')
        return super().form_valid(form)

class UserUpdateView(PaginaAtivaMixin, LoginRequiredMixin, UpdateView):
    model = User
    form_class = CustomUserChangeForm
    template_name = 'claviculario_app/user_form.html'
    success_url = reverse_lazy('user_list')
    pagina_ativa = 'contas'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Editar Usuário: {self.object.username}'
        return context
    def form_valid(self, form):
        messages.success(self.request, 'Usuário atualizado com sucesso!')
        return super().form_valid(form)

class UserDesativarView(LoginRequiredMixin, AccessMixin, UpdateView):
    model = User
    template_name = 'claviculario_app/user_confirm_desativar.html'
    success_url = reverse_lazy('user_list')
    fields = ['is_active'] # O único campo que esta view pode modificar
    
    # Proteção: apenas superusuários podem desativar outros
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return self.handle_no_permission()
        # Não permitir que um superusuário se desative
        if self.get_object() == request.user:
            messages.error(request, 'Você não pode desativar sua própria conta.')
            return redirect('user_list')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        # Em vez de um formulário, nós forçamos o valor para False
        self.object.is_active = False
        self.object.save()
        messages.success(self.request, f"O usuário '{self.object.username}' foi desativado com sucesso.")
        return redirect(self.success_url)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pagina_ativa'] = 'contas'
        return context